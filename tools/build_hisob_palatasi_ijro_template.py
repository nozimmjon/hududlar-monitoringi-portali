from __future__ import annotations

import importlib.util
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList


ROOT = Path(__file__).resolve().parents[1]
GENERATOR = ROOT / "tools" / "generate_andijon_integrated_platform_v7.py"
TEMPLATE_DIR = ROOT / "platform prototypes" / "templates"
TEMPLATE_XLSX = TEMPLATE_DIR / "hisob_palatasi_ijro_tekshiruv_template.xlsx"
TEST_XLSX = TEMPLATE_DIR / "hisob_palatasi_ijro_tekshiruv_test_filled.xlsx"


PERIOD_LABELS = {
    "q1": "I чорак",
    "h1": "II чорак / I ярим йиллик",
    "m9": "III чорак / 9 ой",
    "year": "IV чорак / йил якуни",
}


def load_generator_module():
    spec = importlib.util.spec_from_file_location("andijon_v7_generator", GENERATOR)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {GENERATOR}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_platform_data() -> dict:
    gen = load_generator_module()
    import json

    data = json.loads(gen.DATA_JSON.read_text(encoding="utf-8"))
    gen.normalize_source_data(data)
    gen.enrich_industry_drivers(data)
    data["tasks"] = gen.extract_kafolat_tasks(data)
    return data


def sheet_by_prefix(wb, prefix: str):
    for ws in wb.worksheets:
        if ws.title.startswith(prefix):
            return ws
    raise KeyError(prefix)


def row_style(ws, row: int):
    return [
        {
            "style": copy(ws.cell(row, col)._style),
            "font": copy(ws.cell(row, col).font),
            "fill": copy(ws.cell(row, col).fill),
            "border": copy(ws.cell(row, col).border),
            "alignment": copy(ws.cell(row, col).alignment),
            "number_format": ws.cell(row, col).number_format,
            "protection": copy(ws.cell(row, col).protection),
        }
        for col in range(1, ws.max_column + 1)
    ]


def apply_row_style(ws, row: int, style_row) -> None:
    for col, style in enumerate(style_row, start=1):
        cell = ws.cell(row, col)
        cell._style = copy(style["style"])
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.number_format = style["number_format"]
        cell.protection = copy(style["protection"])


def resize_data_area(ws, data_rows: int, *, template_row: int = 2) -> None:
    target_max = max(1 + data_rows, template_row)
    if ws.max_row > target_max:
        ws.delete_rows(target_max + 1, ws.max_row - target_max)
    elif ws.max_row < target_max:
        ws.insert_rows(ws.max_row + 1, target_max - ws.max_row)

    style = row_style(ws, min(template_row, ws.max_row))
    for row in range(2, target_max + 1):
        apply_row_style(ws, row, style)
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).value = None


def add_list_validation(ws, col: str, row_end: int, formula: str) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{row_end}")


def reset_validations(ws) -> None:
    ws.data_validations = DataValidationList()


def readiness_formula_02(row: int) -> str:
    return (
        f'=IF(COUNTA(L{row}:W{row})=0,"",'
        f'IF(AND(L{row}<>"",M{row}<>"",N{row}<>"",U{row}<>"",V{row}<>"",'
        f'IF(H{row}="Ҳар ой",I{row}<>"",TRUE),'
        f'IF(OR(L{row}="Қайтарилди",L{row}="Рад этилди",M{row}="Бажарилмади",'
        f'M{row}="Муддати кечикди",N{row}<>"Етарли"),'
        f'AND(Q{row}<>"",R{row}<>"",T{row}<>""),TRUE)),"Тайёр","Тўлдирилмаган"))'
    )


def readiness_formula_03(row: int) -> str:
    return (
        f'=IF(COUNTA(H{row}:S{row})=0,"",'
        f'IF(AND(H{row}<>"",I{row}<>"",J{row}<>"",Q{row}<>"",R{row}<>"",'
        f'IF(OR(H{row}="Қайтарилди",H{row}="Рад этилди",I{row}="Бажарилмади",'
        f'I{row}="Муддати кечикди",J{row}<>"Етарли"),'
        f'AND(M{row}<>"",N{row}<>"",P{row}<>""),TRUE)),"Тайёр","Тўлдирилмаган"))'
    )


def readiness_formula_04(row: int) -> str:
    return (
        f'=IF(COUNTA(L{row}:V{row})=0,"",'
        f'IF(AND(L{row}<>"",M{row}<>"",N{row}<>"",T{row}<>"",U{row}<>"",'
        f'IF(OR(L{row}="Қайтарилди",L{row}="Рад этилди",M{row}="Бажарилмади",'
        f'M{row}="Муддати кечикди",N{row}<>"Етарли"),'
        f'AND(P{row}<>"",Q{row}<>"",S{row}<>""),TRUE)),"Тайёр","Тўлдирилмаган"))'
    )


def periodicity_for(task: dict) -> str:
    period = task.get("periodCode")
    if period == "h1":
        return "Ярим йиллик"
    if period == "year":
        return "Йил якуни"
    if period == "m9":
        return "Доимий / давомида"
    return "Бир марталик"


def region_for(task: dict) -> str:
    districts = task.get("districts") or []
    return ", ".join(districts) if districts else "Вилоят"


def update_instruction_sheet(ws, data: dict) -> None:
    meta = data["task_meta"]
    ws["B4"] = f"02_Текширув киритиш: {meta['execution_tasks']} та T-топшириқ бўйича умумий текширув."
    ws["B5"] = "03_Ойлик текширув: ҳозирги Андижон чора-тадбирлар реестрида ойлик T-ID ажратилмаган; зарурат бўлса қўшимча ойлик қаторлар киритилади."
    ws["B6"] = f"04_Туман кесими: {meta['district_targets']} та D-қатор бўйича туман/шаҳар кесимида текширув."
    ws["B10"] = str(data["task_meta"]["source_doc"])
    ws["B11"] = (
        f"{meta['execution_tasks']} та чора-тадбир, "
        f"{meta['kpi_targets']} та KPI reference, "
        f"{meta['district_targets']} та туман кесими"
    )
    ws["B13"] = "2026-05-06"


def update_main_input_sheet(ws, data: dict) -> None:
    tasks = data["tasks"]
    resize_data_area(ws, len(tasks))
    for idx, task in enumerate(tasks, start=2):
        ws.cell(idx, 1, task["id"])
        ws.cell(idx, 2, task.get("sourceNo"))
        ws.cell(idx, 3, task.get("moduleLabel"))
        ws.cell(idx, 4, task.get("kpi"))
        ws.cell(idx, 5, task.get("title"))
        ws.cell(idx, 6, task.get("deadline"))
        ws.cell(idx, 7, PERIOD_LABELS.get(task.get("periodCode"), task.get("period") or ""))
        ws.cell(idx, 8, periodicity_for(task))
        ws.cell(idx, 9, None)
        ws.cell(idx, 10, region_for(task))
        ws.cell(idx, 11, task.get("owner"))
        ws.cell(idx, 24, readiness_formula_02(idx))
    end = len(tasks) + 1
    reset_validations(ws)
    add_list_validation(ws, "L", end, "'05_Луғатлар'!$A$2:$A$6")
    add_list_validation(ws, "M", end, "'05_Луғатлар'!$B$2:$B$6")
    add_list_validation(ws, "N", end, "'05_Луғатлар'!$C$2:$C$4")
    add_list_validation(ws, "Q", end, "'05_Луғатлар'!$D$2:$D$10")
    add_list_validation(ws, "G", end, "'05_Луғатлар'!$E$2:$E$6")
    add_list_validation(ws, "H", end, "'05_Луғатлар'!$F$2:$F$6")
    add_list_validation(ws, "I", end, "'05_Луғатлар'!$G$2:$G$15")
    add_list_validation(ws, "P", end, "'05_Луғатлар'!$H$2:$H$13")


def update_monthly_sheet(ws, rows: int = 24) -> None:
    resize_data_area(ws, rows)
    for idx in range(2, rows + 2):
        ws.cell(idx, 20, readiness_formula_03(idx))
    reset_validations(ws)
    end = rows + 1
    add_list_validation(ws, "H", end, "'05_Луғатлар'!$A$2:$A$6")
    add_list_validation(ws, "I", end, "'05_Луғатлар'!$B$2:$B$6")
    add_list_validation(ws, "J", end, "'05_Луғатлар'!$C$2:$C$4")
    add_list_validation(ws, "M", end, "'05_Луғатлар'!$D$2:$D$10")
    add_list_validation(ws, "F", end, "'05_Луғатлар'!$G$2:$G$14")
    add_list_validation(ws, "L", end, "'05_Луғатлар'!$H$2:$H$13")


def update_district_sheet(ws, data: dict) -> None:
    targets = data.get("kafolat_district_targets", [])
    resize_data_area(ws, len(targets))
    for idx, row in enumerate(targets, start=2):
        ws.cell(idx, 1, row.get("id"))
        ws.cell(idx, 2, row.get("parentTaskId"))
        ws.cell(idx, 3, row.get("sourceNo"))
        ws.cell(idx, 4, row.get("district"))
        ws.cell(idx, 5, row.get("title"))
        ws.cell(idx, 6, PERIOD_LABELS.get(row.get("periodCode"), row.get("period") or ""))
        ws.cell(idx, 7, row.get("deadline"))
        ws.cell(idx, 8, row.get("value"))
        ws.cell(idx, 9, row.get("yearValue"))
        ws.cell(idx, 10, row.get("unit"))
        ws.cell(idx, 11, row.get("note"))
        ws.cell(idx, 23, readiness_formula_04(idx))
    reset_validations(ws)
    end = len(targets) + 1
    add_list_validation(ws, "L", end, "'05_Луғатлар'!$A$2:$A$6")
    add_list_validation(ws, "M", end, "'05_Луғатлар'!$B$2:$B$6")
    add_list_validation(ws, "N", end, "'05_Луғатлар'!$C$2:$C$4")
    add_list_validation(ws, "P", end, "'05_Луғатлар'!$D$2:$D$10")
    add_list_validation(ws, "F", end, "'05_Луғатлар'!$E$2:$E$6")
    add_list_validation(ws, "J", end, "'05_Луғатлар'!$H$2:$H$13")


def registry_rows(data: dict) -> list[dict]:
    rows = []
    for row in data.get("kafolat_kpi_targets", []):
        rows.append({
            "sourceNo": row.get("sourceNo"),
            "type": "KPI",
            "platformId": row.get("platformId") or row.get("id"),
            "module": row.get("moduleLabel"),
            "kpi": row.get("kpi"),
            "title": row.get("title"),
            "deadline": row.get("deadline"),
            "periodCode": row.get("periodCode"),
            "periodicity": periodicity_for(row),
            "districts": region_for(row),
            "owner": row.get("owner"),
        })
    for row in data.get("tasks", []):
        rows.append({
            "sourceNo": row.get("sourceNo"),
            "type": "Чора-тадбирлар",
            "platformId": row.get("id"),
            "module": row.get("moduleLabel"),
            "kpi": row.get("kpi"),
            "title": row.get("title"),
            "deadline": row.get("deadline"),
            "periodCode": row.get("periodCode"),
            "periodicity": periodicity_for(row),
            "districts": region_for(row),
            "owner": row.get("owner"),
        })
    return sorted(rows, key=lambda item: (item.get("sourceNo") or 9999, item.get("platformId") or ""))


def update_registry_sheet(ws, data: dict) -> None:
    rows = registry_rows(data)
    resize_data_area(ws, len(rows))
    for idx, row in enumerate(rows, start=2):
        for col, key in enumerate(
            ["sourceNo", "type", "platformId", "module", "kpi", "title", "deadline", "periodCode", "periodicity", "districts", "owner"],
            start=1,
        ):
            ws.cell(idx, col, row.get(key))


def update_data_model_sheet(ws) -> None:
    ws["E4"] = "T-001...T-054"
    ws["E6"] = "D-01...D-22"
    ws["F6"] = "D-ID + T-ID + туман + давр import key"
    ws["F7"] = "Киритилди/кўриб чиқилмоқда/тасдиқланди/қайтарилди/рад этилди"


def build_template(output: Path, *, fill_samples: bool = False) -> None:
    data = load_platform_data()
    wb = load_workbook(TEMPLATE_XLSX)
    update_instruction_sheet(sheet_by_prefix(wb, "01_"), data)
    update_main_input_sheet(sheet_by_prefix(wb, "02_"), data)
    update_monthly_sheet(sheet_by_prefix(wb, "03_"))
    update_district_sheet(sheet_by_prefix(wb, "04_"), data)
    update_registry_sheet(sheet_by_prefix(wb, "06_"), data)
    update_data_model_sheet(sheet_by_prefix(wb, "07_"))

    if fill_samples:
        ws = sheet_by_prefix(wb, "02_")
        ws["L2"] = "Тасдиқланди"
        ws["M2"] = "Бажарилди"
        ws["N2"] = "Етарли"
        ws["O2"] = 0.5
        ws["P2"] = "трлн сўм"
        ws["R2"] = "НАМУНА: Ҳисоб палатаси тасдиқлаган қатор"
        ws["U2"] = "2026-05-06"
        ws["V2"] = "Ҳисоб палатаси"
        ws["W2"] = "sample-evidence.pdf"

        ws["L32"] = "Қайтарилди"
        ws["M32"] = "Қисман бажарилди"
        ws["N32"] = "Етарли эмас"
        ws["Q32"] = "Далил етарсиз"
        ws["R32"] = "НАМУНА: далил тўлиқ эмас"
        ws["S32"] = "Қўшимча далил киритиш"
        ws["T32"] = "2026-06-15"
        ws["U32"] = "2026-05-06"
        ws["V32"] = "Ҳисоб палатаси"

        dws = sheet_by_prefix(wb, "04_")
        dws["L2"] = "Тасдиқланди"
        dws["M2"] = "Бажарилди"
        dws["N2"] = "Етарли"
        dws["O2"] = 209
        dws["P2"] = "Муаммо йўқ"
        dws["Q2"] = "НАМУНА: туман кесимида тасдиқланган"
        dws["T2"] = "2026-05-06"
        dws["U2"] = "Ҳисоб палатаси"
        dws["V2"] = "sample-district-evidence.pdf"

    wb.save(output)


def main() -> None:
    build_template(TEMPLATE_XLSX)
    shutil.copyfile(TEMPLATE_XLSX, TEST_XLSX)
    build_template(TEST_XLSX, fill_samples=True)
    print(TEMPLATE_XLSX)
    print(TEST_XLSX)


if __name__ == "__main__":
    main()
