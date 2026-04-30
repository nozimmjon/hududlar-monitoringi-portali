from __future__ import annotations

from pathlib import Path

import openpyxl
from docx import Document


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data_source" / "Кафолат хатлар имзога" / "2. Андижон"


def clean(value: object) -> str:
    return str(value).replace("\n", " ").strip()


def print_docx_sections() -> None:
    doc = Document(SRC / "0. Кафолат хати (Андижон).docx")
    paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    print("DOCX paragraphs:", len(paras))
    for idx, text in enumerate(paras, 1):
        marker = ""
        if text.startswith(("I.", "II.", "III.", "IV.", "V.", "VI.")):
            marker = "SECTION"
        elif any(k in text for k in ("ЯҲМ", "Бюджет", "Инвестиция", "Хорижий", "экспорт", "ишсизлик", "камбағал", "инфляц")):
            marker = "KPI"
        if marker:
            print(f"{idx:03d} [{marker}] {text}")


def print_sheet_headers() -> None:
    specs = [
        ("1.1-1.5-жадваллар (макро).xlsx", "1.1. ЯҲМ", range(4, 10)),
        ("1.1-1.5-жадваллар (макро).xlsx", "1.2. Саноат", range(4, 9)),
        ("1.1-1.5-жадваллар (макро).xlsx", "1.3. Ҳудудий саноат", range(4, 8)),
        ("1.1-1.5-жадваллар (макро).xlsx", "1.4. ҚХ", range(4, 9)),
        ("1.1-1.5-жадваллар (макро).xlsx", "1.5. Бозор хизматлари", range(4, 9)),
        ("2.1-2.2-жадваллар (инфляция).xlsx", "1.1. Баланс", range(4, 8)),
        ("2.1-2.2-жадваллар (инфляция).xlsx", "1.2. Омборлар", range(3, 7)),
        ("3-жадвал (бюджет).xlsx", "тушум", range(4, 9)),
        ("4.1-жадвал (бюджет инвестка).xlsx", "2.Анд", range(4, 8)),
        ("4.2-жадвал (инвестициялар).xlsx", "4,2-хорижий инв", range(4, 8)),
        ("5.1-5.2-жадваллар (экспорт).xlsx", "5-жадвал", range(4, 8)),
        ("5.1-5.2-жадваллар (экспорт).xlsx", "02_Анд", range(4, 8)),
        ("5.1-5.2-жадваллар (экспорт).xlsx", "Корхона сони", range(3, 7)),
        ("6-жадвал (бандлик ва камбағаллик даражаси).xlsx", "6. Камбағаллик", range(4, 8)),
    ]
    for file_name, sheet, rows in specs:
        wb = openpyxl.load_workbook(SRC / file_name, read_only=True, data_only=True)
        ws = wb[sheet]
        print(f"\n### {file_name} :: {sheet} ({ws.max_row}x{ws.max_column})")
        for row_idx in rows:
            values = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
            cells = [f"{col}:{clean(value)[:70]}" for col, value in enumerate(values, 1) if value not in (None, "")]
            print(f"row {row_idx}: " + " ; ".join(cells[:36]))


if __name__ == "__main__":
    print_docx_sections()
    print_sheet_headers()
